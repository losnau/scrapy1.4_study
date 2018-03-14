#-*- coding:utf-8
from openpyxl import Workbook
import  xlwt
from scrapy.exporters import BaseItemExporter
class ExcelItemExporter(BaseItemExporter):
    def __init__(self,file,**kwargs):
        self._configure(kwargs)
        self.file=file
        # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
        self.wb2=Workbook()
        self.ws2=self.wb2.active
        self.ws2.title = 'scrapy'
        print('-----------------------------分割线1---------------------------------')
        print(self.file.name)
        print(type(self.file.name))
        print('-=-------------分割线2--------')
        print(self.file)
        print(type(self.file))
        self.row=1
    def finish_exporting(self):
        #self.file 的类型是：<class '_io.BufferedWriter'>
        self.wb2.save(filename=self.file.name)
    def export_item(self, item):
        fields = self._get_serialized_fields(item)
        for col ,v in enumerate(x for _,x in fields):
            print("**************************row="+str(self.row)+'****col='+str(col+1))
            self.ws2.cell(row=self.row,column=col+1,value=v)
        self.row+=1
class ExcelItemExporter2(BaseItemExporter):
    def __init__(self,file,**kwargs):
        self._configure(kwargs)
        self.file=file
        self.wb = xlwt.Workbook()
        self.ws = self.wb.add_sheet('scrapy')
        self.row=0
    def finish_exporting(self):
        self.wb.save(self.file)
    def export_item(self, item):
        fields = self._get_serialized_fields(item)
        for col ,v in enumerate(x for _,x in fields):
            print("**************************row="+str(self.row)+'****col='+str(col+1))
            self.ws.write(self.row,col,v)
        self.row+=1
